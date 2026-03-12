import os
import io
import argparse
import warnings
import numpy as np
import pandas as pd
import joblib
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

warnings.filterwarnings("ignore")

# ── Paths ─────────────────────────────────────────────────────────────────────
DATA_PATH  = "ml/data_sets.xlsx"
MODEL_PATH = "ml/models/sales_model.pkl"
META_PATH  = "ml/models/metadata.pkl"
OUTPUT_DIR = "output_data"
ENGLISH_SEASON_BY_MONTH = {
    1: "Winter",
    2: "Winter",
    3: "Spring",
    4: "Spring",
    5: "Summer",
    6: "Summer",
    7: "Monsoon",
    8: "Monsoon",
    9: "Autumn",
    10: "Autumn",
    11: "Pre-Winter",
    12: "Pre-Winter",
}


def load_model_and_meta():
    """Load the saved XGBoost model and all metadata."""
    if not os.path.exists(MODEL_PATH) or not os.path.exists(META_PATH):
        raise FileNotFoundError("Model not found. Please run train.ipynb first.")

    model = joblib.load(MODEL_PATH)
    meta  = joblib.load(META_PATH)
    return model, meta


def merge_into_master(df_new: pd.DataFrame) -> tuple:
    df_master = pd.read_excel(DATA_PATH, sheet_name="Monthly_Data")
    df_master["Discount Band"] = df_master["Discount Band"].fillna("None")

    new_year  = int(df_new["Year"].max())
    new_month = int(df_new[df_new["Year"] == new_year]["Month Number"].max())

    already_exists = (
        (df_master["Year"] == new_year) &
        (df_master["Month Number"] == new_month)
    ).any()

    if already_exists:
        label = pd.Timestamp(f"{new_year}-{new_month:02d}-01").strftime("%B %Y")
        print(f"Note: {label} already exists in data_sets.xlsx — skipping merge.")
        return df_master, False

    df_combined = (
        pd.concat([df_master, df_new], ignore_index=True)
        .sort_values(["Product", "Year", "Month Number"])
        .reset_index(drop=True)
    )

    with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        df_combined.to_excel(w, sheet_name="Monthly_Data", index=False)

    label = pd.Timestamp(f"{new_year}-{new_month:02d}-01").strftime("%B %Y")
    print(f"Merged    : {label} added to data_sets.xlsx → {len(df_combined):,} rows total")
    return df_combined, True


def predict_next_month(df_raw: pd.DataFrame, model, meta: dict) -> dict:
    FEATURE_COLS  = meta["feature_cols"]
    PROD_MAP      = meta["prod_map"]
    SEG_ENC       = meta["seg_enc"]
    ACTIVE_MONTHS = meta["active_months"]
    PRODUCT_SPECS = meta["product_specs"]
    PRODUCT_DISC  = meta["product_disc"]
    BASE_YEAR     = meta["base_year"]
    SEASON_NAME_RAW = meta["season_name"]
    SEASON_ENC_RAW  = meta["season_enc"]
    ALL_PRODUCTS  = meta["all_products"]
    DISC_ENC      = meta["disc_enc"]
    DISC_RATE     = meta["disc_rate"]

    # Auto-detect next month from latest data
    latest_year   = int(df_raw["Year"].max())
    latest_month  = int(df_raw[df_raw["Year"] == latest_year]["Month Number"].max())
    latest_date   = pd.Timestamp(f"{latest_year}-{latest_month:02d}-01")
    next_date     = latest_date + pd.DateOffset(months=1)
    predict_year  = int(next_date.year)
    predict_month = int(next_date.month)
    season_label  = ENGLISH_SEASON_BY_MONTH.get(predict_month, "Unknown")

    print(f"Latest data : {latest_date.strftime('%B %Y')}")
    print(f"Predicting  : {next_date.strftime('%B %Y')}  ({season_label})")

    def predict_one_product(product: str) -> dict:
        hist = df_raw[df_raw["Product"] == product].copy()
        hist["Date"] = pd.to_datetime(
            hist["Year"].astype(str) + "-" +
            hist["Month Number"].astype(str).str.zfill(2) + "-01"
        )
        hist = hist.sort_values("Date").reset_index(drop=True)

        spec    = PRODUCT_SPECS[product]
        mfg     = spec["Manufacturing Price"]
        sprice  = spec["Sale Price"]
        seg     = spec["Segment"]
        disc    = PRODUCT_DISC[product]
        td      = pd.Timestamp(f"{predict_year}-{predict_month:02d}-01")
        quarter = (predict_month - 1) // 3 + 1

        same_mo = hist[hist["Month Number"] == predict_month]
        units   = same_mo["Units Sold"].mean() if not same_mo.empty else hist["Units Sold"].mean()

        def get_lag(lag):
            cutoff = td - pd.DateOffset(months=lag)
            row    = hist[hist["Date"] == cutoff]
            if not row.empty:
                return row["Sales"].values[0]
            sm = hist[hist["Month Number"] == cutoff.month]
            return sm["Sales"].mean() if not sm.empty else hist["Sales"].mean()

        sv = np.array(hist["Sales"].values.astype(float))
        def smean(n): return float(np.mean(sv[-n:])) if len(sv) >= n else float(np.mean(sv))
        def sstd(n):  return float(np.std(sv[-n:]))  if len(sv) >= n else 0.0

        dr = DISC_RATE.get(disc, 0)
        eg = units * sprice

        features = {
            "Units Sold":          units,
            "Manufacturing Price": mfg,
            "Sale Price":          sprice,
            "Gross Sales":         eg,
            "Discounts":           eg * dr,
            "COGS":                units * mfg,
            "Month":               predict_month,
            "Year_num":            predict_year,
            "Quarter":             quarter,
            "Month_sin":           float(np.sin(2 * np.pi * predict_month / 12)),
            "Month_cos":           float(np.cos(2 * np.pi * predict_month / 12)),
            "Qtr_sin":             float(np.sin(2 * np.pi * quarter / 4)),
            "Qtr_cos":             float(np.cos(2 * np.pi * quarter / 4)),
            "Season_enc":          SEASON_ENC_RAW.get(SEASON_NAME_RAW.get(predict_month), 0),
            "Year_trend":          predict_year - BASE_YEAR,
            "Segment_enc":         SEG_ENC.get(seg, 0),
            "Product_enc":         PROD_MAP.get(product, 0),
            "DiscBand_enc":        DISC_ENC.get(disc, 0),
            "Discount_Rate":       dr,
            "Revenue_per_Unit":    sprice * (1 - dr),
            "Cost_per_Unit":       mfg,
            "Price_to_ManufCost":  sprice / mfg if mfg else 0,
            "Sales_lag1":          get_lag(1),
            "Sales_lag2":          get_lag(2),
            "Sales_lag3":          get_lag(3),
            "Sales_lag6":          get_lag(6),
            "Sales_lag12":         get_lag(12),
            "Sales_roll3_mean":    smean(3),
            "Sales_roll3_std":     sstd(3),
            "Sales_roll6_mean":    smean(6),
            "Sales_roll6_std":     sstd(6),
        }

        X = pd.DataFrame([features])[FEATURE_COLS]
        predicted = round(float(model.predict(X)[0]), 2)
        predicted_profit = round(float(predicted - (units * mfg)), 2)

        return {
            "product":          product,
            "segment":          seg,
            "season":           season_label,
            "predicted_units":  round(float(units), 0),
            "predicted_sales":  predicted,
            "predicted_profit": predicted_profit,
        }

    # Predict all active products for next month
    active  = [p for p in ALL_PRODUCTS if predict_month in ACTIVE_MONTHS[p]]
    results = [predict_one_product(p) for p in active]
    results.sort(key=lambda x: x["predicted_sales"], reverse=True)

    total = round(sum(r["predicted_sales"] for r in results), 2)

    # Save output Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_filename = f"predicted sales-report-{next_date.strftime('%Y-%m')}.xlsx"
    out_path     = os.path.join(OUTPUT_DIR, out_filename)
    prediction_frame = pd.DataFrame(results).rename(columns={
        "product":         "Product",
        "segment":         "Segment",
        "season":          "Season",
        "predicted_units": "Predicted_Units",
        "predicted_sales": "Predicted_Sales",
        "predicted_profit": "Predicted_Profit",
    })[["Product", "Segment", "Season", "Predicted_Units", "Predicted_Sales", "Predicted_Profit"]]
    prediction_frame.to_excel(out_path, sheet_name="Predictions", index=False)

    if not prediction_frame.empty:
        workbook = load_workbook(out_path)
        prediction_sheet = workbook["Predictions"]
        charts_sheet = workbook.create_sheet("Charts")

        def add_pie_chart(title: str, value_column: int, anchor: str) -> None:
            labels = Reference(
                prediction_sheet,
                min_col=1,
                min_row=2,
                max_row=len(prediction_frame) + 1,
            )
            data = Reference(
                prediction_sheet,
                min_col=value_column,
                min_row=1,
                max_row=len(prediction_frame) + 1,
            )
            chart = PieChart()
            chart.title = title
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 7
            chart.width = 9
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showLeaderLines = True
            charts_sheet.add_chart(chart, anchor)

        add_pie_chart("Predicted Unit Sales Share", 4, "A1")
        add_pie_chart("Predicted Sales Share", 5, "J1")
        add_pie_chart("Predicted Profit Share", 6, "A20")

        workbook.save(out_path)

    print(f"Products : {len(active)}")
    print(f"Total    : {total:,.2f}")
    print(f"Saved    : {out_path}")

    return {
        "latest_data_month":     latest_date.strftime("%B %Y"),
        "predicting_month":      next_date.strftime("%B %Y"),
        "season":                season_label,
        "active_products":       len(active),
        "total_predicted_sales": total,
        "predictions":           results,
        "output_file":           out_filename,
    }


def run_from_file(new_data_file: str) -> dict:
    if not os.path.exists(new_data_file):
        raise FileNotFoundError(f"File not found: {new_data_file}")

    model, meta = load_model_and_meta()

    df_new = pd.read_excel(new_data_file)
    df_new["Discount Band"] = df_new["Discount Band"].fillna("None")

    required = ["Product", "Segment", "Discount Band", "Units Sold",
                "Manufacturing Price", "Sale Price", "Gross Sales",
                "Discounts", "Sales", "COGS", "Profit", "Month Number", "Year"]
    missing = [c for c in required if c not in df_new.columns]
    if missing:
        raise ValueError(f"File is missing columns: {missing}")

    print(f"New file  : {new_data_file}  ({len(df_new):,} rows)")

    df_raw, was_merged = merge_into_master(df_new)
    result = predict_next_month(df_raw, model, meta)
    result["merged"] = was_merged
    return result


def run_from_latest() -> dict:
    model, meta = load_model_and_meta()
    df_raw = pd.read_excel(DATA_PATH, sheet_name="Monthly_Data")
    df_raw["Discount Band"] = df_raw["Discount Band"].fillna("None")
    return predict_next_month(df_raw, model, meta)


# ── CLI entry point ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sales Forecast Predictor")
    group  = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file",   type=str, help="Path to new sales-report-YYYY-MM.xlsx")
    group.add_argument("--latest", action="store_true", help="Predict from existing data_sets.xlsx")
    args = parser.parse_args()

    if args.file:
        run_from_file(args.file)
    else:
        run_from_latest()
